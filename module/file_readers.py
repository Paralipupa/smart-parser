import abc
import csv
import os
import logging
import xlrd
from openpyxl import load_workbook
from typing import List

logger = logging.getLogger(__name__)


def rchop(s, sub):
    return s[: -len(sub)] if s.endswith(sub) else s


class DataFile(abc.ABC):
    def __init__(self, fname):
        self._fname = fname
        self._sheet = None

    def __iter__(self):
        return self

    def __next__(self):
        return ""


class CsvReaderWrapper:
    """ обёртка класса csv.reader для добавления своих атрибутов """
    def __init__(self, reader, title):
        self._reader = reader
        self.title = title

    def __iter__(self):
        return self

    def __next__(self):
        return next(self._reader)


class CsvFile(DataFile):
    def __init__(self, fname):
        super(CsvFile, self).__init__(fname)
        self._freader = open(fname, "r", encoding="cp1251")
        self._reader = CsvReaderWrapper(
            csv.reader(self._freader, delimiter=";", quotechar="|"),
            title="Лист1",
        )
        self._rows = iter(self._reader)
        self._line_num = 0

    def set_config(
        self, page_indexes: List[int] = [], number_columns: int = 150
    ) -> bool:
        self._page_current = 0
        self._columns = range(number_columns)
        self._page_indexes = [0]
        return self._page_current

    def get_sheets(self):
        return [self._reader]

    def get_sheet(self):
        # Возвращаем текущий "лист" (всегда один для CSV)
        if self._page_current == 0:
            self._page_current += 1
            return 0
        return None

    def __next__(self):
        try:
            return next(self._rows)
        except StopIteration:
            self._freader.close()
            raise StopIteration

    def __del__(self):
        if not self._freader.closed:
            self._freader.close()


class XlsFile(DataFile):
    def __init__(self, fname):
        super(XlsFile, self).__init__(fname)
        self._wb = xlrd.open_workbook(fname)

    def set_config(
        self, page_indexes: List[int] = [], number_columns: int = 150
    ) -> bool:
        self._page_current = 0
        self._columns = range(number_columns)
        if page_indexes:
            self._page_indexes = page_indexes
        else:
            self._page_indexes = range(len(self._wb.sheets()))
        return self._page_current

    def get_sheets(self):
        return self._wb.sheets()

    def get_sheet(self) -> object:
        try:
            if self._page_current < len(self._page_indexes):
                self._sheet = self._wb.sheets()[self._page_indexes[self._page_current]]
                self._rows = (
                    self._sheet.row(index) for index in range(self._sheet.nrows)
                )
                self._page_current += 1
                return self._page_current - 1
            return None
        except Exception as ex:
            logger.exception("getSheet")

    @staticmethod
    def get_cell_text(cell):
        if cell.ctype == 2:
            return rchop(str(cell.value), ".0")
        return str(cell.value)

    def get_row(self, row):
        index = 0
        for cell in row:
            if index in self._columns:
                yield XlsFile.get_cell_text(cell)
            index = index + 1

    def __next__(self):
        for row in self._rows:
            return list(self.get_row(row))
        raise StopIteration

    def __del__(self):
        self._wb.release_resources()
        del self._wb


class XlsxFile(DataFile):
    def __init__(self, fname):
        super(XlsxFile, self).__init__(fname)
        try:
            self._wb = load_workbook(filename=fname, data_only=True)
        except:
            self._wb = load_workbook(filename=fname, read_only=True, data_only=True)

    def set_config(
        self, page_indexes: List[int] = [], number_columns: int = 150
    ) -> bool:
        self._page_current = 0
        self._columns = range(number_columns)
        if page_indexes:
            self._page_indexes = page_indexes
        else:
            self._page_indexes = range(len(self._wb.worksheets))
        return self._page_current

    def get_sheets(self):
        return self._wb.worksheets

    def get_sheet(self) -> object:
        try:
            if self._page_current < len(self._page_indexes):
                self._sheet = self._wb.worksheets[
                    self._page_indexes[self._page_current]
                ]
                self._cursor = self._sheet.iter_rows()
                self._page_current += 1
                return self._page_current - 1
            return None
        except Exception as ex:
            logger.exception("getSheet")

    @staticmethod
    def get_cell_text(cell):
        return str(cell.value) if cell.value != None else ""

    def get_row(self, row):
        i = 0
        for cell in row:
            if i in self._columns:
                yield XlsxFile.get_cell_text(cell)
            i += 1

    def __next__(self):
        return list(self.get_row(next(self._cursor)))

    def __del__(self):
        self._wb.close()
        del self._wb

    def get_index(self, cell):
        try:
            return cell.column
        except AttributeError:
            return -1


def get_file_reader(fname):
    """Get class for reading file as iterable"""
    _, file_extension = os.path.splitext(fname)
    if file_extension == ".xls":
        return XlsFile
    if file_extension == ".xlsx":
        # return XlsFile
        return XlsxFile
    if file_extension == ".csv":
        return CsvFile
    raise Exception("Unknown file type")
